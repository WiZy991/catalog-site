"""
Дополнительные view для админки каталога.
"""
import csv
import io
import os
from django.shortcuts import render, redirect
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib import messages
from django.db import transaction
from django.http import HttpResponse
from django.core.files.base import ContentFile
from django.db.models import Max
from django.utils.text import get_valid_filename
import openpyxl
import xlrd

from .forms import BulkImageUploadForm, BulkProductImportForm, QuickProductForm
from .services import (
    process_bulk_images, 
    process_bulk_import, 
    parse_product_name,
    get_category_for_product,
)
from .models import Product, ProductImage


@staff_member_required
def bulk_image_upload(request):
    """Массовая загрузка изображений."""
    if request.method == 'POST':
        form = BulkImageUploadForm(request.POST, request.FILES)
        if form.is_valid():
            files = request.FILES.getlist('images')
            create_products = form.cleaned_data['create_products']
            apply_single_to_all = form.cleaned_data.get('apply_single_to_all', False)

            if apply_single_to_all:
                if not files:
                    messages.error(request, 'Не выбрано изображение для массового применения.')
                    return redirect(request.path)

                source_file = files[0]
                source_bytes = source_file.read()
                source_name = source_file.name or 'shared.jpg'
                base_name, ext = os.path.splitext(source_name)
                ext = ext or '.jpg'

                active_products = list(
                    Product.objects.filter(is_active=True)
                    .only('id', 'article', 'brand')
                    .order_by('id')
                )
                updated_count = 0
                BATCH_SIZE = 50

                for i in range(0, len(active_products), BATCH_SIZE):
                    batch = active_products[i:i + BATCH_SIZE]
                    with transaction.atomic():
                        for product in batch:
                            order = (product.images.aggregate(max_order=Max('order')).get('max_order') or 0) + 1
                            file_name = get_valid_filename(f'{product.article or product.pk}_shared{ext}')
                            ProductImage.objects.create(
                                product=product,
                                image=ContentFile(source_bytes, name=file_name),
                                is_main=not product.images.exists(),
                                order=order,
                                alt=f'Фото {product.article} {product.brand}'.strip()
                            )
                            updated_count += 1

                if len(files) > 1:
                    messages.info(request, 'Использовано только первое изображение из списка для применения ко всем товарам.')

                messages.success(request, f'✅ Одно фото добавлено в {updated_count} активных товаров.')
                return redirect('admin:catalog_product_changelist')
            
            # Собираем изображения
            images = []
            for f in files:
                images.append((f.name, f.read()))
            
            # Обрабатываем
            stats = process_bulk_images(images, create_products=create_products)
            
            # Показываем результаты
            messages.success(
                request, 
                f'✅ Загружено изображений: {stats["matched"]} из {stats["total"]}'
            )
            
            if stats['created_products']:
                messages.info(
                    request,
                    f'📦 Создано новых товаров: {stats["created_products"]}'
                )
            
            if stats.get('duplicates', 0) > 0:
                messages.info(
                    request,
                    f'ℹ️ Пропущено дубликатов: {stats["duplicates"]}'
                )
                # Показываем список дубликатов
                if stats.get('duplicate_files', [])[:10]:
                    files_list = ', '.join(stats['duplicate_files'][:10])
                    if len(stats['duplicate_files']) > 10:
                        files_list += f' и ещё {len(stats["duplicate_files"]) - 10}...'
                    messages.info(request, f'Дубликаты: {files_list}')
            
            if stats['not_matched']:
                messages.warning(
                    request,
                    f'⚠️ Не удалось привязать: {stats["not_matched"]} файлов'
                )
                # Показываем список непривязанных
                if stats['not_matched_files'][:10]:
                    files_list = ', '.join(stats['not_matched_files'][:10])
                    if len(stats['not_matched_files']) > 10:
                        files_list += f' и ещё {len(stats["not_matched_files"]) - 10}...'
                    messages.warning(request, f'Файлы: {files_list}')
            
            return redirect('admin:catalog_product_changelist')
    else:
        form = BulkImageUploadForm()
    
    return render(request, 'admin/catalog/bulk_image_upload.html', {
        'form': form,
        'title': 'Массовая загрузка изображений',
    })


@staff_member_required
def bulk_product_import(request):
    """Массовый импорт товаров из файла."""
    if request.method == 'POST':
        form = BulkProductImportForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            auto_category = form.cleaned_data['auto_category']
            auto_brand = form.cleaned_data['auto_brand']
            
            # Читаем файл
            data_rows = []
            filename = file.name.lower()
            
            try:
                if filename.endswith('.xml'):
                    # XML файл (CommerceML 2 формат)
                    import xml.etree.ElementTree as ET
                    from .commerceml_views import parse_commerceml_product, process_product_from_commerceml
                    
                    file.seek(0)  # Сбрасываем позицию файла
                    content = file.read()
                    
                    # Парсим XML
                    try:
                        root = ET.fromstring(content)
                    except ET.ParseError as e:
                        raise Exception(f'Ошибка парсинга XML: {str(e)}')
                    
                    # Автоматически определяем namespace из корневого элемента
                    # Формат: {namespace}ИмяЭлемента
                    root_tag = root.tag
                    if root_tag.startswith('{'):
                        # Извлекаем namespace из тега
                        namespace = root_tag[1:root_tag.index('}')]
                        root_tag_name = root_tag[root_tag.index('}')+1:]
                    else:
                        namespace = None
                        root_tag_name = root_tag
                    
                    # Определяем namespace CommerceML (поддерживаем разные варианты)
                    namespaces = {}
                    if namespace:
                        # Используем найденный namespace
                        namespaces['cml'] = namespace
                        namespaces[''] = namespace
                    else:
                        # Стандартные namespace CommerceML
                        namespaces['cml'] = 'http://v8.1c.ru/8.3/commerceml'
                        namespaces[''] = 'http://v8.1c.ru/8.3/commerceml'
                    
                    # Также добавляем альтернативный namespace (urn:1C.ru:commerceml_2)
                    namespaces['cml2'] = 'urn:1C.ru:commerceml_2'
                    
                    # Ищем каталог товаров (пробуем разные варианты)
                    catalog = None
                    
                    # Вариант 1: С namespace
                    if namespace:
                        catalog = root.find(f'.//{{{namespace}}}Каталог') or root.find(f'.//{{{namespace}}}catalog')
                    
                    # Вариант 2: Стандартные варианты
                    if catalog is None:
                        catalog = (
                            root.find('.//catalog', namespaces) or 
                            root.find('.//Каталог', namespaces) or
                            root.find('.//catalog') or 
                            root.find('.//Каталог')
                        )
                    
                    if catalog is None:
                        # Выводим структуру для диагностики
                        error_msg = f'Каталог не найден в XML файле. '
                        error_msg += f'Корневой элемент: {root_tag_name}, namespace: {namespace or "нет"}. '
                        error_msg += f'Дочерние элементы: {[child.tag for child in root[:5]]}'
                        raise Exception(error_msg)
                    
                    # Извлекаем товары (пробуем разные варианты)
                    products_elements = []
                    
                    # Вариант 1: С namespace
                    if namespace:
                        products_elements = (
                            catalog.findall(f'.//{{{namespace}}}Товары/{{{namespace}}}Товар') or
                            catalog.findall(f'.//{{{namespace}}}Товар')
                        )
                    
                    # Вариант 2: Стандартные варианты
                    if not products_elements:
                        products_elements = (
                            catalog.findall('.//catalog:Товары/catalog:Товар', namespaces) or
                            catalog.findall('.//Товары/Товар') or
                            catalog.findall('.//catalog:Товар', namespaces) or
                            catalog.findall('.//Товар')
                        )
                    
                    if not products_elements:
                        error_msg = 'Товары не найдены в XML файле. '
                        error_msg += f'Найдены элементы в каталоге: {[child.tag for child in catalog[:5]]}'
                        raise Exception(error_msg)
                    
                    # Преобразуем товары в формат для process_bulk_import
                    for product_elem in products_elements:
                        product_data = parse_commerceml_product(product_elem, namespaces, root)
                        if product_data and product_data.get('name'):
                            # Преобразуем в формат, ожидаемый process_bulk_import
                            row = {
                                'name': product_data.get('name', ''),
                                'article': product_data.get('article', '') or product_data.get('sku', ''),
                            }
                            
                            # External ID для поиска/обновления товаров
                            if product_data.get('external_id'):
                                row['external_id'] = product_data.get('external_id')
                            
                            # Бренд
                            if product_data.get('brand'):
                                row['brand'] = product_data.get('brand')
                            
                            # Категория
                            if product_data.get('category_name'):
                                row['category_name'] = product_data.get('category_name')
                            
                            # Цена
                            if product_data.get('price'):
                                price_val = product_data.get('price')
                                row['price'] = str(price_val).replace('.', ',')
                                row['price_num'] = float(price_val)
                            
                            # Остаток
                            if 'stock' in product_data:
                                qty_val = product_data.get('stock', 0)
                                row['quantity'] = str(qty_val)
                                row['quantity_num'] = int(qty_val)
                                row['остаток_num'] = int(qty_val)
                            
                            # Описание
                            if product_data.get('description'):
                                row['description'] = product_data.get('description')
                            
                            # Применимость
                            if product_data.get('applicability'):
                                row['applicability'] = product_data.get('applicability')
                            
                            # Кросс-номера
                            if product_data.get('cross_numbers'):
                                row['cross_numbers'] = product_data.get('cross_numbers')
                            
                            # Обрабатываем характеристики
                            if product_data.get('characteristics'):
                                char_parts = []
                                for char in product_data['characteristics']:
                                    if isinstance(char, dict):
                                        char_parts.append(f"{char.get('name', '')}: {char.get('value', '')}")
                                if char_parts:
                                    row['characteristics'] = '\n'.join(char_parts)
                            
                            data_rows.append(row)
                    
                    if not data_rows:
                        raise Exception('Не удалось извлечь товары из XML файла.')
                        
                elif filename.endswith('.csv'):
                    # CSV файл
                    content = file.read().decode('utf-8-sig')
                    reader = csv.DictReader(io.StringIO(content), delimiter=';')
                    for row in reader:
                        # Нормализуем ключи и обрабатываем значения
                        normalized = {}
                        for key, value in row.items():
                            if key:
                                key_normalized = key.lower().strip()
                                # Обрабатываем числовые значения
                                if value and value.strip():
                                    # Пытаемся определить, является ли значение числом
                                    try:
                                        # Убираем пробелы и проверяем формат
                                        value_clean = value.replace(' ', '').replace('\xa0', '').replace(',', '.')
                                        num_value = float(value_clean)
                                        # Если это число, сохраняем и строковое, и числовое значение
                                        normalized[key_normalized] = value.strip()
                                        # Сохраняем числовое значение для цен и остатков
                                        if 'цена' in key_normalized or 'price' in key_normalized:
                                            normalized['price_num'] = num_value
                                        elif 'остаток' in key_normalized or 'quantity' in key_normalized or 'склад' in key_normalized:
                                            normalized['quantity_num'] = int(num_value)
                                            normalized['остаток_num'] = int(num_value)
                                    except (ValueError, TypeError):
                                        # Не число, сохраняем как строку
                                        normalized[key_normalized] = value.strip()
                                else:
                                    normalized[key_normalized] = value.strip() if value else ''
                        data_rows.append(normalized)
                        
                elif filename.endswith(('.xls', '.xlsx')):
                    # Excel файл
                    file.seek(0)  # Сбрасываем позицию файла
                    
                    # Определяем формат файла: старый .xls (бинарный) или новый .xlsx (zip)
                    is_old_xls = filename.endswith('.xls') and not filename.endswith('.xlsx')
                    
                    if is_old_xls:
                        # Старый формат .xls - используем xlrd
                        try:
                            file_content = file.read()
                            wb = xlrd.open_workbook(file_contents=file_content)
                            ws = wb.sheet_by_index(0)
                            
                            # Ищем строку с заголовками
                            header_row_index = 0
                            headers = []
                            
                            # Проверяем первые 15 строк на наличие заголовков
                            for row_num in range(min(15, ws.nrows)):
                                row_values = []
                                for col_num in range(ws.ncols):
                                    cell_value = ws.cell_value(row_num, col_num)
                                    if cell_value:
                                        row_values.append(str(cell_value).strip())
                                    else:
                                        row_values.append('')
                                
                                # Объединяем все значения строки для проверки
                                row_text = ' '.join(row_values).lower()
                                
                                # Проверяем, есть ли в этой строке ключевые слова заголовков
                                header_keywords = ['артикул', 'номенклатура', 'наименование', 'цена', 'остаток', 'склад', 'розничная', 'фарпост']
                                keyword_count = sum(1 for keyword in header_keywords if keyword in row_text)
                                
                                # Если найдено минимум 2 ключевых слова, считаем это строкой заголовков
                                if keyword_count >= 2:
                                    header_row_index = row_num
                                    headers = row_values
                                    break
                            
                            # Если заголовки не найдены, берем первую строку
                            if not headers:
                                headers = [str(ws.cell_value(0, col_num) or '').strip() for col_num in range(ws.ncols)]
                            
                            # Читаем данные начиная со строки после заголовков
                            for row_num in range(header_row_index + 1, ws.nrows):
                                row_data = {}
                                for col_num in range(min(len(headers), ws.ncols)):
                                    if headers[col_num]:
                                        header_key = headers[col_num].lower().strip()
                                        cell = ws.cell(row_num, col_num)
                                        value = cell.value
                                        
                                        # Обрабатываем значение в зависимости от типа
                                        if value is None or value == '':
                                            value = ''
                                        elif cell.ctype == xlrd.XL_CELL_NUMBER:
                                            # Числовое значение
                                            if isinstance(value, float) and value.is_integer():
                                                value_str = str(int(value))
                                            else:
                                                value_str = str(value)
                                            value_str = value_str.replace('.', ',')
                                            row_data[header_key] = value_str
                                            # Сохраняем числовое значение
                                            if 'цена' in header_key or 'price' in header_key:
                                                row_data[header_key + '_num'] = value
                                            elif 'остаток' in header_key or 'quantity' in header_key or 'склад' in header_key:
                                                row_data[header_key + '_num'] = int(value) if isinstance(value, float) and value.is_integer() else int(value)
                                            continue
                                        else:
                                            value = str(value).strip()
                                        
                                        row_data[header_key] = value
                                
                                # Пропускаем полностью пустые строки
                                has_data = any(
                                    str(v).strip() for v in row_data.values() 
                                    if v is not None and str(v).strip() and not str(v).endswith('_num')
                                )
                                if has_data:
                                    data_rows.append(row_data)
                                    
                        except xlrd.biffh.XLRDError as e:
                            raise Exception(f'Ошибка чтения Excel файла (старый формат .xls): {str(e)}. Убедитесь, что файл не поврежден.')
                        except Exception as e:
                            raise Exception(f'Ошибка при обработке Excel файла: {str(e)}')
                    else:
                        # Новый формат .xlsx - используем openpyxl
                        try:
                            wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
                            ws = wb.active
                            
                            # Ищем строку с заголовками (может быть не в первой строке)
                            # Ищем строку, которая содержит ключевые слова заголовков
                            header_row_index = 1
                            headers = []
                            
                            # Проверяем первые 15 строк на наличие заголовков (увеличено для надежности)
                            for row_num in range(1, min(16, ws.max_row + 1)):
                                row = list(ws.iter_rows(min_row=row_num, max_row=row_num))[0]
                                row_values = []
                                for cell in row:
                                    if cell.value is not None:
                                        # Объединяем многострочные заголовки в одну строку
                                        cell_value = str(cell.value).strip()
                                        row_values.append(cell_value)
                                    else:
                                        row_values.append('')
                                
                                # Объединяем все значения строки для проверки
                                row_text = ' '.join(row_values).lower()
                                
                                # Проверяем, есть ли в этой строке ключевые слова заголовков
                                header_keywords = ['артикул', 'номенклатура', 'наименование', 'цена', 'остаток', 'склад', 'розничная', 'фарпост']
                                keyword_count = sum(1 for keyword in header_keywords if keyword in row_text)
                                
                                # Если найдено минимум 2 ключевых слова, считаем это строкой заголовков
                                if keyword_count >= 2:
                                    header_row_index = row_num
                                    # Сохраняем оригинальные заголовки (не в нижнем регистре, чтобы сохранить формат)
                                    headers = row_values
                                    break
                            
                            # Если заголовки не найдены, берем первую строку
                            if not headers:
                                row = list(ws.iter_rows(min_row=1, max_row=1))[0]
                                headers = [str(cell.value or '').strip() for cell in row]
                            
                            # Читаем данные начиная со строки после заголовков
                            for row_num, row in enumerate(ws.iter_rows(min_row=header_row_index + 1, values_only=False), start=header_row_index + 1):
                                row_data = {}
                                for i, cell in enumerate(row):
                                    if i < len(headers) and headers[i]:
                                        header_key = headers[i].lower().strip()
                                        value = cell.value
                                        
                                        # Обрабатываем значение в зависимости от типа
                                        if value is None:
                                            value = ''
                                        elif isinstance(value, (int, float)):
                                            # Для чисел сохраняем как строку, чтобы сохранить форматирование
                                            # Но также сохраняем числовое значение для правильной обработки
                                            if isinstance(value, float) and value.is_integer():
                                                value_str = str(int(value))
                                            else:
                                                value_str = str(value)
                                            # Заменяем точку на запятую для соответствия формату клиента
                                            value_str = value_str.replace('.', ',')
                                            row_data[header_key] = value_str
                                            # Также сохраняем оригинальное значение для числовых полей
                                            if 'цена' in header_key or 'price' in header_key:
                                                row_data[header_key + '_num'] = value
                                                # Если это оптовая цена, сохраняем отдельно
                                                if 'опт' in header_key or 'wholesale' in header_key:
                                                    row_data['wholesale_price_num'] = value
                                            elif 'остаток' in header_key or 'quantity' in header_key or 'склад' in header_key:
                                                row_data[header_key + '_num'] = int(value) if isinstance(value, float) and value.is_integer() else int(value)
                                            continue
                                        else:
                                            value = str(value).strip()
                                        
                                        row_data[header_key] = value
                                
                                # Пропускаем полностью пустые строки
                                has_data = any(
                                    str(v).strip() for v in row_data.values() 
                                    if v is not None and str(v).strip() and not str(v).endswith('_num')
                                )
                                if has_data:
                                    data_rows.append(row_data)
                            
                            wb.close()
                        except openpyxl.utils.exceptions.InvalidFileException as e:
                            raise Exception(f'Файл не является корректным Excel файлом (.xlsx). Возможно, файл поврежден или имеет неправильный формат. Ошибка: {str(e)}')
                        except Exception as e:
                            raise Exception(f'Ошибка при чтении Excel файла: {str(e)}')
                
                # Маппинг колонок для формата прайс-листа клиента и 1С
                def normalize_key(key):
                    key_lower = key.lower().strip()
                    
                    # Название товара
                    if key_lower in ['наименование для печати', 'рабочее наименование']:
                        return 'name'
                    if 'наименование' in key_lower and 'печат' in key_lower:
                        return 'name'
                    if 'номенклатура' in key_lower and 'характеристика' in key_lower:
                        return 'name'
                    
                    # Категория (номенклатура без характеристики)
                    if key_lower == 'номенклатура':
                        return 'category_name'
                    
                    # Артикул1 - основной артикул бренда (кросс-номер)
                    if key_lower == 'артикул1' or key_lower == 'артикул 1':
                        return 'article'
                    
                    # Артикул2 - OEM номер
                    if key_lower == 'артикул2' or key_lower == 'артикул 2':
                        return 'oem_number'
                    
                    # Обычный артикул (если нет разделения на 1 и 2)
                    if key_lower == 'артикул' or key_lower == 'article':
                        return 'article'
                    
                    # Марка/Бренд
                    if key_lower in ['марка', 'бренд', 'brand', 'производитель']:
                        return 'brand'
                    
                    # Двигатель → применимость
                    if key_lower in ['двигатель', 'engine', 'мотор']:
                        return 'engine'
                    
                    # Кузов → применимость
                    if key_lower in ['кузов', 'body', 'кузова']:
                        return 'body'
                    
                    # Модель → применимость
                    if key_lower in ['модель', 'model', 'модели']:
                        return 'model'
                    
                    # Размер → характеристики (может быть вольтаж, материал и т.д.)
                    if key_lower in ['размер', 'size', 'габариты']:
                        return 'size'
                    
                    # Позиционирование
                    if key_lower in ['l-r', 'лево-право', 'сторона']:
                        return 'side'
                    if key_lower in ['f-r', 'перед-зад', 'позиция']:
                        return 'position'
                    if key_lower in ['u-d', 'верх-низ', 'направление']:
                        return 'direction'
                    
                    # Другие поля
                    if key_lower in ['год', 'year', 'годы']:
                        return 'year'
                    if key_lower in ['цвет', 'color']:
                        return 'color'
                    if key_lower in ['примечание', 'note', 'комментарий']:
                        return 'note'
                    if key_lower in ['новый', 'состояние', 'condition']:
                        return 'condition'
                    
                    # Цена
                    if 'цена' in key_lower or 'розничная' in key_lower or 'farpost' in key_lower or 'руб' in key_lower or key_lower == 'price':
                        return 'price'
                    
                    # Оптовая цена
                    if 'опт' in key_lower and 'цена' in key_lower:
                        return 'wholesale_price'
                    
                    # Остаток
                    if 'остаток' in key_lower or 'склад' in key_lower or key_lower == 'quantity':
                        return 'quantity'
                    
                    # Стандартные поля
                    if key_lower in ['name', 'brand', 'category', 'description', 'applicability', 'cross_numbers', 'availability']:
                        return key_lower
                    
                    return key_lower
                
                # Применяем маппинг колонок
                mapped_rows = []
                for row in data_rows:
                    mapped = {}
                    
                    # Сначала обрабатываем обычные ключи (не _num)
                    for key, value in row.items():
                        # Пропускаем служебные ключи с _num - их обработаем отдельно
                        if key.endswith('_num'):
                            continue
                        
                        mapped_key = normalize_key(key)
                        
                        # Если несколько колонок маппятся на один ключ, берем первую непустую
                        if mapped_key in mapped and mapped[mapped_key]:
                            continue
                        
                        mapped[mapped_key] = value
                    
                    # Теперь обрабатываем числовые значения (_num)
                    for orig_key in row.keys():
                        if orig_key.endswith('_num'):
                            num_value = row[orig_key]
                            base_key = orig_key.replace('_num', '')
                            base_mapped_key = normalize_key(base_key)
                            
                            if base_mapped_key == 'price' or 'цена' in base_key.lower() or 'price' in base_key.lower():
                                mapped['price_num'] = num_value
                            elif base_mapped_key == 'quantity' or 'остаток' in base_key.lower() or 'quantity' in base_key.lower() or 'склад' in base_key.lower():
                                mapped['quantity_num'] = num_value
                                mapped['остаток_num'] = num_value  # Дублируем для надежности
                    
                    mapped_rows.append(mapped)
                
                # Импортируем
                stats = process_bulk_import(
                    mapped_rows, 
                    auto_category=auto_category,
                    auto_brand=auto_brand
                )
                
                messages.success(
                    request,
                    f'✅ Импорт завершён! Создано: {stats["created"]}, обновлено: {stats["updated"]}'
                )
                
                if stats['errors']:
                    messages.warning(
                        request,
                        f'⚠️ Ошибок: {stats["errors"]}'
                    )
                    for error in stats['error_details'][:5]:
                        messages.error(request, error)
                
                return redirect('admin:catalog_product_changelist')
                
            except Exception as e:
                messages.error(request, f'Ошибка при чтении файла: {str(e)}')
    else:
        form = BulkProductImportForm()
    
    return render(request, 'admin/catalog/bulk_product_import.html', {
        'form': form,
        'title': 'Массовый импорт товаров',
    })


@staff_member_required
def quick_add_product(request):
    """Быстрое добавление товара с автоматическим заполнением."""
    if request.method == 'POST':
        form = QuickProductForm(request.POST, request.FILES)
        if form.is_valid():
            name = form.cleaned_data['name']
            price = form.cleaned_data.get('price') or 0
            image = form.cleaned_data.get('image')
            
            # Парсим название
            parsed = parse_product_name(name)
            
            # Определяем категорию автоматически
            category = get_category_for_product(name)
            
            # Создаём товар в ОСНОВНОМ каталоге
            product = Product.objects.create(
                name=name,
                article=parsed['article'] or '',
                brand=parsed['brand'] or '',
                category=category,
                price=price,
                catalog_type='retail',  # ОСНОВНОЙ КАТАЛОГ!
                is_active=True,
            )
            
            # Загружаем изображение (проверяем на дубликаты)
            if image:
                # Проверяем, нет ли уже изображений у товара
                if not product.images.exists():
                    ProductImage.objects.create(
                        product=product,
                        image=image,
                        is_main=True,
                    )
                else:
                    # Если изображения уже есть, не создаем дубликат
                    messages.warning(
                        request,
                        f'⚠️ У товара "{product.name}" уже есть изображения. Новое изображение не добавлено.'
                    )
            
            # Сообщения
            success_msg = f'✅ Товар "{product.name}" создан!'
            info_messages = []
            if parsed['brand']:
                info_messages.append(f'🏭 Бренд определён: {parsed["brand"]}')
            if parsed['article']:
                info_messages.append(f'📦 Артикул определён: {parsed["article"]}')
            if parsed['category']:
                info_messages.append(f'📁 Категория: {parsed["category"]}')
            
            # Очищаем форму и показываем сообщения
            form = QuickProductForm()
            messages.success(request, success_msg)
            for msg in info_messages:
                messages.info(request, msg)
            
            # Рендерим страницу без редиректа, чтобы избежать дублирования
            return render(request, 'admin/catalog/quick_add_product.html', {
                'form': form,
                'title': 'Быстрое добавление товара',
            })
    else:
        form = QuickProductForm()
    
    return render(request, 'admin/catalog/quick_add_product.html', {
        'form': form,
        'title': 'Быстрое добавление товара',
    })


@staff_member_required
def download_import_template(request):
    """Скачать шаблон для импорта."""
    response = HttpResponse(
        content_type='text/csv; charset=utf-8-sig'
    )
    response['Content-Disposition'] = 'attachment; filename="import_template.csv"'
    
    writer = csv.writer(response, delimiter=';')
    writer.writerow([
        'Название', 'Артикул', 'Бренд', 'Цена', 'Категория',
        'Описание', 'Применимость', 'Кросс-номера', 'Наличие', 'Состояние', 'Farpost URL'
    ])
    # Пример данных
    writer.writerow([
        'Стартер Isuzu 10PD1 24V', 'ME220745', 'Isuzu', '15000',
        'Стартеры', 'Новый оригинальный стартер', 'Isuzu Forward, Isuzu Giga',
        '1-81100-141-0, 0-23000-1670', 'in_stock', 'new', ''
    ])
    
    return response

