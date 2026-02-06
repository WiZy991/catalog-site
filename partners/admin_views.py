"""
Дополнительные view для админки партнёрского раздела.
Массовый импорт товаров с оптовыми ценами.
"""
import csv
import io
from django.shortcuts import render, redirect
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib import messages
from django.http import HttpResponse
import openpyxl
import xlrd

from catalog.models import Product, Category
from catalog.services import get_category_for_product


@staff_member_required
def bulk_wholesale_import(request):
    """Массовый импорт товаров с оптовыми ценами для партнёрского раздела."""
    if request.method == 'POST':
        file = request.FILES.get('file')
        if not file:
            messages.error(request, 'Выберите файл для импорта')
            return redirect('admin_bulk_wholesale_import')
        
        auto_category = request.POST.get('auto_category') == 'on'
        auto_brand = request.POST.get('auto_brand') == 'on'
        update_existing = request.POST.get('update_existing') == 'on'
        
        # Читаем файл
        data_rows = []
        filename = file.name.lower()
        
        try:
            if filename.endswith('.csv'):
                # CSV файл
                content = file.read().decode('utf-8-sig')
                reader = csv.DictReader(io.StringIO(content), delimiter=';')
                for row in reader:
                    normalized = {}
                    for key, value in row.items():
                        if key:
                            key_normalized = key.lower().strip()
                            if value and value.strip():
                                try:
                                    value_clean = value.replace(' ', '').replace('\xa0', '').replace(',', '.')
                                    num_value = float(value_clean)
                                    normalized[key_normalized] = value.strip()
                                    if 'цена' in key_normalized or 'price' in key_normalized:
                                        if 'опт' in key_normalized or 'wholesale' in key_normalized:
                                            normalized['wholesale_price_num'] = num_value
                                        else:
                                            normalized['price_num'] = num_value
                                    elif 'остаток' in key_normalized or 'quantity' in key_normalized or 'склад' in key_normalized:
                                        normalized['quantity_num'] = int(num_value)
                                except (ValueError, TypeError):
                                    normalized[key_normalized] = value.strip()
                            else:
                                normalized[key_normalized] = value.strip() if value else ''
                    data_rows.append(normalized)
                    
            elif filename.endswith(('.xls', '.xlsx')):
                # Excel файл
                file.seek(0)
                is_old_xls = filename.endswith('.xls') and not filename.endswith('.xlsx')
                
                if is_old_xls:
                    # Старый формат .xls
                    file_content = file.read()
                    wb = xlrd.open_workbook(file_contents=file_content)
                    ws = wb.sheet_by_index(0)
                    
                    header_row_index = 0
                    headers = []
                    
                    for row_num in range(min(15, ws.nrows)):
                        row_values = []
                        for col_num in range(ws.ncols):
                            cell_value = ws.cell_value(row_num, col_num)
                            row_values.append(str(cell_value).strip() if cell_value else '')
                        
                        row_text = ' '.join(row_values).lower()
                        header_keywords = ['артикул', 'номенклатура', 'наименование', 'цена', 'остаток', 'склад', 'опт', 'wholesale']
                        keyword_count = sum(1 for keyword in header_keywords if keyword in row_text)
                        
                        if keyword_count >= 2:
                            header_row_index = row_num
                            headers = row_values
                            break
                    
                    if not headers:
                        headers = [str(ws.cell_value(0, col_num) or '').strip() for col_num in range(ws.ncols)]
                    
                    for row_num in range(header_row_index + 1, ws.nrows):
                        row_data = {}
                        for col_num in range(min(len(headers), ws.ncols)):
                            if headers[col_num]:
                                header_key = headers[col_num].lower().strip()
                                cell = ws.cell(row_num, col_num)
                                value = cell.value
                                
                                if value is None or value == '':
                                    value = ''
                                elif cell.ctype == xlrd.XL_CELL_NUMBER:
                                    if isinstance(value, float) and value.is_integer():
                                        value_str = str(int(value))
                                    else:
                                        value_str = str(value)
                                    value_str = value_str.replace('.', ',')
                                    row_data[header_key] = value_str
                                    if 'цена' in header_key or 'price' in header_key:
                                        if 'опт' in header_key or 'wholesale' in header_key:
                                            row_data['wholesale_price_num'] = value
                                        else:
                                            row_data['price_num'] = value
                                    elif 'остаток' in header_key or 'quantity' in header_key or 'склад' in header_key:
                                        row_data['quantity_num'] = int(value) if isinstance(value, float) and value.is_integer() else int(value)
                                    continue
                                else:
                                    value = str(value).strip()
                                
                                row_data[header_key] = value
                        
                        has_data = any(str(v).strip() for v in row_data.values() if v and not str(v).endswith('_num'))
                        if has_data:
                            data_rows.append(row_data)
                else:
                    # Новый формат .xlsx
                    wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
                    ws = wb.active
                    
                    header_row_index = 1
                    headers = []
                    
                    for row_num in range(1, min(16, ws.max_row + 1)):
                        row = list(ws.iter_rows(min_row=row_num, max_row=row_num))[0]
                        row_values = [str(cell.value).strip() if cell.value else '' for cell in row]
                        
                        row_text = ' '.join(row_values).lower()
                        header_keywords = ['артикул', 'номенклатура', 'наименование', 'цена', 'остаток', 'склад', 'опт', 'wholesale']
                        keyword_count = sum(1 for keyword in header_keywords if keyword in row_text)
                        
                        if keyword_count >= 2:
                            header_row_index = row_num
                            headers = row_values
                            break
                    
                    if not headers:
                        row = list(ws.iter_rows(min_row=1, max_row=1))[0]
                        headers = [str(cell.value or '').strip() for cell in row]
                    
                    for row_num, row in enumerate(ws.iter_rows(min_row=header_row_index + 1, values_only=False), start=header_row_index + 1):
                        row_data = {}
                        for i, cell in enumerate(row):
                            if i < len(headers) and headers[i]:
                                header_key = headers[i].lower().strip()
                                value = cell.value
                                
                                if value is None:
                                    value = ''
                                elif isinstance(value, (int, float)):
                                    if isinstance(value, float) and value.is_integer():
                                        value_str = str(int(value))
                                    else:
                                        value_str = str(value)
                                    value_str = value_str.replace('.', ',')
                                    row_data[header_key] = value_str
                                    if 'цена' in header_key or 'price' in header_key:
                                        if 'опт' in header_key or 'wholesale' in header_key:
                                            row_data['wholesale_price_num'] = value
                                        else:
                                            row_data['price_num'] = value
                                    elif 'остаток' in header_key or 'quantity' in header_key or 'склад' in header_key:
                                        row_data['quantity_num'] = int(value) if isinstance(value, float) and value.is_integer() else int(value)
                                    continue
                                else:
                                    value = str(value).strip()
                                
                                row_data[header_key] = value
                        
                        has_data = any(str(v).strip() for v in row_data.values() if v and not str(v).endswith('_num'))
                        if has_data:
                            data_rows.append(row_data)
                    
                    wb.close()
            
            # Маппинг колонок
            def normalize_key(key):
                key_lower = key.lower().strip()
                
                if 'номенклатура' in key_lower or 'наименование' in key_lower or 'характеристика' in key_lower or 'печать' in key_lower:
                    return 'name'
                if 'артикул' in key_lower or key_lower == 'article':
                    return 'article'
                # Оптовая цена (приоритет)
                if 'опт' in key_lower or 'wholesale' in key_lower:
                    return 'wholesale_price'
                # Розничная цена
                if 'цена' in key_lower or 'розничная' in key_lower or 'farpost' in key_lower or key_lower == 'price':
                    return 'price'
                if 'остаток' in key_lower or 'склад' in key_lower or key_lower == 'quantity':
                    return 'quantity'
                if key_lower in ['name', 'brand', 'category', 'description', 'applicability', 'cross_numbers', 'condition', 'availability']:
                    return key_lower
                return key_lower
            
            # Применяем маппинг
            mapped_rows = []
            for row in data_rows:
                mapped = {}
                
                for key, value in row.items():
                    if key.endswith('_num'):
                        continue
                    mapped_key = normalize_key(key)
                    if mapped_key in mapped and mapped[mapped_key]:
                        continue
                    mapped[mapped_key] = value
                
                # Числовые значения
                for orig_key in row.keys():
                    if orig_key.endswith('_num'):
                        num_value = row[orig_key]
                        if 'wholesale' in orig_key or 'опт' in orig_key:
                            mapped['wholesale_price_num'] = num_value
                        elif 'price' in orig_key or 'цена' in orig_key:
                            mapped['price_num'] = num_value
                        elif 'quantity' in orig_key or 'остаток' in orig_key or 'склад' in orig_key:
                            mapped['quantity_num'] = num_value
                
                mapped_rows.append(mapped)
            
            # Импортируем товары с оптовыми ценами
            from catalog.services import process_bulk_import_wholesale
            stats = process_bulk_import_wholesale(
                mapped_rows, 
                auto_category=auto_category,
                auto_brand=auto_brand,
                update_existing=update_existing
            )
            
            messages.success(
                request,
                f'✅ Импорт завершён! Создано: {stats["created"]}, обновлено: {stats["updated"]}'
            )
            
            if stats['errors']:
                messages.warning(request, f'⚠️ Ошибок: {stats["errors"]}')
                for error in stats['error_details'][:5]:
                    messages.error(request, error)
            
            return redirect('admin:partners_wholesaleproduct_changelist')
            
        except Exception as e:
            messages.error(request, f'Ошибка при чтении файла: {str(e)}')
    
    return render(request, 'admin/partners/bulk_wholesale_import.html', {
        'title': 'Массовый импорт оптовых товаров',
    })


@staff_member_required
def download_wholesale_template(request):
    """Скачать шаблон для импорта оптовых товаров."""
    response = HttpResponse(
        content_type='text/csv; charset=utf-8-sig'
    )
    response['Content-Disposition'] = 'attachment; filename="wholesale_import_template.csv"'
    
    writer = csv.writer(response, delimiter=';')
    writer.writerow([
        'Артикул', 'Название', 'Бренд', 'Розничная цена', 'Оптовая цена', 'Остаток'
    ])
    # Пример данных
    writer.writerow([
        'ME220745', 'Стартер Isuzu 10PD1 24V', 'Isuzu', '15000', '12000', '5'
    ])
    writer.writerow([
        '1-81100-141-0', 'Генератор Toyota 2C 12V', 'Toyota', '18000', '15500', '3'
    ])
    
    return response
