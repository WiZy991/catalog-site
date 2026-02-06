from django.contrib import admin
from django.contrib import messages
from django.contrib.auth.models import User
from django.utils.html import format_html
from django.utils import timezone
from django.core.mail import send_mail
from django.conf import settings
import secrets
import string

from .models import PartnerRequest, Partner, PartnerSettings, PartnerOrder, PartnerOrderItem
from catalog.models import WholesaleProduct, ProductImage
from django.utils.html import format_html
from import_export import resources, fields
from import_export.admin import ImportExportModelAdmin
from import_export.widgets import ForeignKeyWidget


class PartnerRequestAdmin(admin.ModelAdmin):
    """Админка для заявок партнёров."""
    list_display = [
        'full_name', 'email', 'phone', 'city', 'status_badge', 'created_at', 'processed_at'
    ]
    list_filter = ['status', 'created_at', 'city']
    search_fields = ['full_name', 'email', 'phone', 'city', 'comment']
    readonly_fields = ['created_at', 'updated_at', 'processed_at', 'processed_by', 'partner']
    date_hierarchy = 'created_at'
    ordering = ['-created_at']
    actions = ['approve_requests', 'reject_requests', 'create_partner_for_approved']
    
    fieldsets = (
        ('Контактные данные', {
            'fields': ('full_name', 'phone', 'email', 'city', 'comment')
        }),
        ('Статус', {
            'fields': ('status', 'admin_comment', 'partner')
        }),
        ('Информация', {
            'fields': ('created_at', 'updated_at', 'processed_at', 'processed_by'),
            'classes': ('collapse',)
        }),
    )
    
    def status_badge(self, obj):
        colors = {
            'pending': '#ffc107',
            'approved': '#28a745',
            'rejected': '#dc3545',
        }
        return format_html(
            '<span style="background: {}; color: #fff; padding: 4px 8px; border-radius: 4px; font-size: 11px;">{}</span>',
            colors.get(obj.status, '#6c757d'),
            obj.get_status_display()
        )
    status_badge.short_description = 'Статус'
    
    def approve_requests(self, request, queryset):
        """Одобрить выбранные заявки и создать партнёров."""
        approved_count = 0
        for req in queryset.filter(status='pending'):
            partner = self.create_partner_from_request(req, request.user)
            if partner:
                approved_count += 1
        
        if approved_count:
            self.message_user(
                request, 
                f'Одобрено заявок: {approved_count}. Партнёрам отправлены данные для входа.',
                messages.SUCCESS
            )
        else:
            self.message_user(
                request,
                'Нет заявок для одобрения (возможно, они уже обработаны).',
                messages.WARNING
            )
    approve_requests.short_description = 'Одобрить выбранные заявки'
    
    def reject_requests(self, request, queryset):
        """Отклонить выбранные заявки."""
        count = queryset.filter(status='pending').update(
            status='rejected',
            processed_at=timezone.now(),
            processed_by=request.user
        )
        self.message_user(request, f'Отклонено заявок: {count}', messages.SUCCESS)
    reject_requests.short_description = 'Отклонить выбранные заявки'
    
    def create_partner_for_approved(self, request, queryset):
        """Создать партнёра для уже одобренных заявок."""
        created_count = 0
        for req in queryset.filter(status='approved', partner__isnull=True):
            partner = self.create_partner_from_request(req, request.user)
            if partner:
                created_count += 1
        
        if created_count:
            self.message_user(
                request,
                f'Создано партнёров: {created_count}. Партнёрам отправлены данные для входа.',
                messages.SUCCESS
            )
        else:
            self.message_user(
                request,
                'Нет заявок для обработки (возможно, партнёры уже созданы или заявки не одобрены).',
                messages.WARNING
            )
    create_partner_for_approved.short_description = 'Создать партнёра для одобренных заявок'
    
    def create_partner_from_request(self, req, admin_user):
        """Создать партнёра из заявки."""
        # Если партнёр уже создан и связан с этой заявкой, возвращаем его
        if req.partner:
            return req.partner
        
        # Генерируем пароль
        password = self.generate_password()
        
        # Создаём пользователя
        username = req.email
        user = None
        
        if User.objects.filter(username=username).exists():
            # Пользователь уже существует
            user = User.objects.get(username=username)
            if hasattr(user, 'partner_profile'):
                # Уже есть профиль партнёра
                partner = user.partner_profile
                
                # Проверяем, не связан ли этот партнёр с другой заявкой
                existing_request = PartnerRequest.objects.filter(partner=partner).exclude(pk=req.pk).first()
                if existing_request:
                    # Партнёр уже связан с другой заявкой
                    # Обновляем только статус текущей заявки, но не связываем с партнёром
                    # (чтобы не нарушать ограничение OneToOneField)
                    req.status = 'approved'
                    req.processed_at = timezone.now()
                    req.processed_by = admin_user
                    # Сохраняем БЕЗ partner
                    req.save(update_fields=['status', 'processed_at', 'processed_by', 'updated_at'])
                    # Отправляем новый пароль
                    user.set_password(password)
                    user.save()
                    self.send_credentials_email(req.email, req.full_name, password)
                    return partner
                
                # Партнёр не связан с другой заявкой - можем связать
                # Сначала обновляем статус без partner
                req.status = 'approved'
                req.processed_at = timezone.now()
                req.processed_by = admin_user
                req.save(update_fields=['status', 'processed_at', 'processed_by', 'updated_at'])
                
                # Затем связываем через update, чтобы избежать конфликта
                PartnerRequest.objects.filter(pk=req.pk).update(partner=partner)
                req.refresh_from_db()
                
                # Отправляем новый пароль
                user.set_password(password)
                user.save()
                self.send_credentials_email(req.email, req.full_name, password)
                return partner
        else:
            # Создаём нового пользователя
            user = User.objects.create_user(
                username=username,
                email=req.email,
                password=password,
                first_name=req.full_name.split()[0] if req.full_name else '',
                last_name=' '.join(req.full_name.split()[1:]) if req.full_name else '',
            )
        
        # Создаём профиль партнёра
        partner = Partner.objects.create(
            user=user,
            full_name=req.full_name,
            phone=req.phone,
            city=req.city,
        )
        
        # Обновляем заявку - сначала обновляем статус и другие поля
        req.status = 'approved'
        if not req.processed_at:
            req.processed_at = timezone.now()
        if not req.processed_by:
            req.processed_by = admin_user
        
        # Сохраняем заявку БЕЗ partner, чтобы избежать конфликта
        req.save(update_fields=['status', 'processed_at', 'processed_by', 'updated_at'])
        
        # Теперь связываем партнёра с заявкой
        # Используем update для избежания конфликта
        PartnerRequest.objects.filter(pk=req.pk).update(partner=partner)
        req.refresh_from_db()
        
        # Отправляем email с данными для входа
        self.send_credentials_email(req.email, req.full_name, password)
        
        return partner
    
    def generate_password(self, length=12):
        """Генерация случайного пароля."""
        alphabet = string.ascii_letters + string.digits
        return ''.join(secrets.choice(alphabet) for _ in range(length))
    
    def send_credentials_email(self, email, name, password):
        """Отправка данных для входа партнёру."""
        try:
            subject = f'Доступ к партнёрскому разделу - {settings.SITE_NAME}'
            message = f'''
Здравствуйте, {name}!

Ваша заявка на партнёрство одобрена.

Данные для входа в партнёрский раздел:

Логин: {email}
Пароль: {password}

Ссылка для входа: {settings.SITE_NAME}/partners/login/

С уважением,
Команда {settings.SITE_NAME}
'''
            send_mail(
                subject=subject,
                message=message,
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[email],
                fail_silently=True,
            )
        except Exception:
            pass


class PartnerAdmin(admin.ModelAdmin):
    """Админка для партнёров."""
    list_display = [
        'full_name', 'company_name', 'email', 'phone', 'city', 
        'discount_percent', 'is_active_badge', 'created_at'
    ]
    list_filter = ['is_active', 'city', 'created_at']
    search_fields = ['full_name', 'company_name', 'user__email', 'phone', 'inn']
    readonly_fields = ['created_at', 'updated_at', 'last_login', 'user']
    list_editable = ['discount_percent']
    ordering = ['-created_at']
    actions = ['activate_partners', 'deactivate_partners']
    
    fieldsets = (
        ('Связь с пользователем', {
            'fields': ('user',)
        }),
        ('Контактные данные', {
            'fields': ('full_name', 'company_name', 'phone', 'city')
        }),
        ('Реквизиты', {
            'fields': ('inn', 'kpp', 'legal_address'),
            'classes': ('collapse',)
        }),
        ('Настройки', {
            'fields': ('is_active', 'discount_percent')
        }),
        ('Информация', {
            'fields': ('created_at', 'updated_at', 'last_login'),
            'classes': ('collapse',)
        }),
    )
    
    def email(self, obj):
        return obj.user.email
    email.short_description = 'Email'
    
    def is_active_badge(self, obj):
        if obj.is_active:
            return format_html(
                '<span style="background: #28a745; color: #fff; padding: 4px 8px; border-radius: 4px; font-size: 11px;">Активен</span>'
            )
        return format_html(
            '<span style="background: #dc3545; color: #fff; padding: 4px 8px; border-radius: 4px; font-size: 11px;">Неактивен</span>'
        )
    is_active_badge.short_description = 'Статус'
    
    def activate_partners(self, request, queryset):
        count = queryset.update(is_active=True)
        self.message_user(request, f'Активировано партнёров: {count}', messages.SUCCESS)
    activate_partners.short_description = 'Активировать выбранных партнёров'
    
    def deactivate_partners(self, request, queryset):
        count = queryset.update(is_active=False)
        self.message_user(request, f'Деактивировано партнёров: {count}', messages.SUCCESS)
    deactivate_partners.short_description = 'Деактивировать выбранных партнёров'


class PartnerSettingsAdmin(admin.ModelAdmin):
    """Админка для настроек партнёрского раздела."""
    list_display = ['__str__', 'manager_email', 'updated_at']
    
    fieldsets = (
        ('Email', {
            'fields': ('manager_email',)
        }),
        ('Информация', {
            'fields': ('updated_at',),
            'classes': ('collapse',)
        }),
    )
    readonly_fields = ['updated_at']


class PartnerOrderItemInline(admin.TabularInline):
    """Инлайн для товаров в заказе партнёра."""
    model = PartnerOrderItem
    extra = 0
    fields = ['product', 'quantity', 'price', 'get_total']
    readonly_fields = ['get_total']
    
    def get_total(self, obj):
        if obj.pk:
            return f'{obj.get_total():.2f} ₽'
        return '-'
    get_total.short_description = 'Сумма'


@admin.register(PartnerOrder)
class PartnerOrderAdmin(admin.ModelAdmin):
    """Админка для заказов партнёров."""
    list_display = [
        'id', 'partner', 'status_badge', 'created_at', 'total_price', 'total_quantity'
    ]
    list_filter = ['status', 'created_at']
    search_fields = ['partner__full_name', 'partner__company_name', 'id']
    readonly_fields = ['created_at', 'updated_at']
    inlines = [PartnerOrderItemInline]
    list_display_links = ['id']
    list_per_page = 50
    actions = ['export_orders_xls', 'export_order_xls']
    
    def export_orders_xls(self, request, queryset):
        """Экспорт выбранных заказов в XLS."""
        from django.http import HttpResponse
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from datetime import datetime
        
        wb = Workbook()
        ws = wb.active
        ws.title = 'Заказы партнёров'
        
        # Заголовки (товары слева, контакты справа)
        headers = ['№ заказа', 'Дата', 'Статус', 'Товар', 'Артикул', 'Бренд', 'Количество', 'Цена', 'Сумма', 'Партнёр', 'Компания', 'Телефон', 'Email', 'Комментарий']
        ws.append(headers)
        
        # Стили для заголовков
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Данные (товары слева, контакты справа)
        for order in queryset:
            for item in order.items.all():
                ws.append([
                    order.id,
                    order.created_at.strftime('%d.%m.%Y %H:%M'),
                    order.get_status_display(),
                    item.product.name,
                    item.product.article or '',
                    item.product.brand or '',
                    item.quantity,
                    float(item.price),
                    float(item.get_total()),
                    order.partner.full_name,
                    order.partner.company_name or '',
                    order.partner.phone,
                    order.partner.email,
                    order.comment or '',
                ])
        
        # Автоширина колонок
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="partner_orders_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        wb.save(response)
        return response
    export_orders_xls.short_description = 'Экспортировать выбранные заказы в XLS'
    
    def export_order_xls(self, request, queryset):
        """Экспорт одного заказа в XLS."""
        if queryset.count() != 1:
            self.message_user(request, 'Выберите ровно один заказ для экспорта', messages.WARNING)
            return
        
        order = queryset.first()
        from django.http import HttpResponse
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from datetime import datetime
        
        wb = Workbook()
        ws = wb.active
        ws.title = f'Заказ #{order.id}'
        
        # Заголовки
        headers = ['№ заказа', 'Дата', 'Статус', 'Товар', 'Артикул', 'Бренд', 'Количество', 'Цена', 'Сумма', 'Партнёр', 'Компания', 'Телефон', 'Email', 'Комментарий']
        ws.append(headers)
        
        # Стили для заголовков
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Данные
        for item in order.items.all():
            ws.append([
                order.id,
                order.created_at.strftime('%d.%m.%Y %H:%M'),
                order.get_status_display(),
                item.product.name,
                item.product.article or '',
                item.product.brand or '',
                item.quantity,
                float(item.price),
                float(item.get_total()),
                order.partner.full_name,
                order.partner.company_name or '',
                order.partner.phone,
                order.partner.email,
                order.comment or '',
            ])
        
        # Автоширина колонок
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="partner_order_{order.id}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        wb.save(response)
        return response
    export_order_xls.short_description = 'Экспортировать заказ в XLS'
    
    def status_badge(self, obj):
        colors = {
            'draft': '#6c757d',
            'pending': '#ffc107',
            'processing': '#17a2b8',
            'completed': '#28a745',
            'cancelled': '#dc3545',
        }
        return format_html(
            '<span style="background: {}; color: #fff; padding: 4px 8px; border-radius: 4px; font-size: 11px;">{}</span>',
            colors.get(obj.status, '#6c757d'),
            obj.get_status_display()
        )
    status_badge.short_description = 'Статус'
    
    def total_price(self, obj):
        return f'{obj.get_total_price():.2f} ₽'
    total_price.short_description = 'Сумма'
    
    def total_quantity(self, obj):
        return obj.get_total_quantity()
    total_quantity.short_description = 'Количество'


class WholesaleProductImageInline(admin.TabularInline):
    """Инлайн для изображений оптового товара."""
    model = ProductImage
    extra = 1
    fields = ['image', 'is_main', 'alt', 'order', 'image_preview']
    readonly_fields = ['image_preview']

    def image_preview(self, obj):
        if obj.image:
            return format_html('<img src="{}" style="max-height: 50px;"/>', obj.image.url)
        return '-'
    image_preview.short_description = 'Превью'


class WholesaleProductResource(resources.ModelResource):
    """Ресурс для импорта/экспорта оптовых товаров."""
    from catalog.models import Category
    
    category = fields.Field(
        column_name='category',
        attribute='category',
        widget=ForeignKeyWidget(Category, 'slug')
    )

    class Meta:
        model = WholesaleProduct
        fields = (
            'id', 'name', 'article', 'brand', 'category', 'price', 'wholesale_price', 'old_price',
            'condition', 'availability', 'quantity', 'short_description', 
            'description', 'characteristics', 'applicability', 'cross_numbers',
            'is_active', 'is_featured', 'catalog_type'
        )
        export_order = fields
        import_id_fields = ['article']
        skip_unchanged = True
        report_skipped = True
    
    def before_import_row(self, row, **kwargs):
        """Устанавливаем catalog_type='wholesale' для всех импортируемых товаров."""
        row['catalog_type'] = 'wholesale'
        
        # Нормализация данных
        if 'name' in row and row['name']:
            row['name'] = str(row['name']).strip()
        if 'article' in row and row['article']:
            row['article'] = str(row['article']).strip()
        if 'brand' in row and row['brand']:
            row['brand'] = str(row['brand']).strip()
        
        return super().before_import_row(row, **kwargs)


@admin.register(WholesaleProduct)
class WholesaleProductAdmin(ImportExportModelAdmin, admin.ModelAdmin):
    """Админка для оптовых товаров - импорт для партнёрского каталога."""
    resource_class = WholesaleProductResource
    list_display = [
        'image_preview', 'name', 'article', 'brand', 'category', 
        'price', 'wholesale_price', 'quantity', 'availability', 'is_active'
    ]
    list_display_links = ['name']
    list_filter = ['is_active', 'availability', 'category', 'brand']
    list_editable = ['price', 'wholesale_price', 'availability', 'is_active']
    search_fields = ['name', 'article', 'brand', 'cross_numbers', 'applicability']
    prepopulated_fields = {'slug': ('name',)}
    autocomplete_fields = ['category']
    inlines = [WholesaleProductImageInline]
    actions = ['make_active', 'make_inactive']
    list_per_page = 50
    save_on_top = True
    
    # Кнопка массового импорта
    change_list_template = 'admin/partners/wholesaleproduct_changelist.html'
    
    def get_urls(self):
        from django.urls import path
        urls = super().get_urls()
        # Добавляем URL для массового импорта
        info = self.model._meta.app_label, self.model._meta.model_name
        custom_urls = [
            path('bulk-import/', self.admin_site.admin_view(self.bulk_import_redirect), name='%s_%s_bulk_import' % info),
        ]
        return custom_urls + urls
    
    def bulk_import_redirect(self, request):
        from django.shortcuts import redirect
        return redirect('admin_bulk_wholesale_import')
    
    def make_active(self, request, queryset):
        queryset.update(is_active=True)
        self.message_user(request, f'Активировано товаров: {queryset.count()}', messages.SUCCESS)
    make_active.short_description = 'Сделать активными'
    
    def make_inactive(self, request, queryset):
        queryset.update(is_active=False)
        self.message_user(request, f'Деактивировано товаров: {queryset.count()}', messages.SUCCESS)
    make_inactive.short_description = 'Сделать неактивными'
    
    fieldsets = (
        ('Основное', {
            'fields': ('name', 'slug', 'article', 'brand', 'category')
        }),
        ('Цены', {
            'fields': ('price', 'wholesale_price', 'old_price'),
            'description': 'Розничная цена для основного сайта. Оптовая цена видна только партнёрам.'
        }),
        ('Наличие', {
            'fields': ('condition', 'availability', 'quantity')
        }),
        ('Описание', {
            'fields': ('short_description', 'description', 'characteristics')
        }),
        ('Применимость', {
            'fields': ('applicability', 'cross_numbers')
        }),
        ('SEO', {
            'fields': ('meta_title', 'meta_description'),
            'classes': ('collapse',)
        }),
        ('Настройки', {
            'fields': ('is_active', 'is_featured', 'catalog_type')
        }),
    )
    
    def save_model(self, request, obj, form, change):
        """Автоматически устанавливаем catalog_type='wholesale' для оптовых товаров."""
        if not obj.catalog_type:
            obj.catalog_type = 'wholesale'
        super().save_model(request, obj, form, change)

    def image_preview(self, obj):
        img = obj.get_main_image()
        if img and img.image:
            return format_html('<img src="{}" style="max-height: 50px;"/>', img.image.url)
        return '-'
    image_preview.short_description = 'Фото'
    
    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        extra_context['bulk_import_url'] = '/admin/partners/bulk-import/'
        return super().changelist_view(request, extra_context=extra_context)
    
    def get_queryset(self, request):
        # Показываем ТОЛЬКО товары из партнёрского каталога
        return super().get_queryset(request).filter(catalog_type='wholesale')


admin.site.register(PartnerRequest, PartnerRequestAdmin)
admin.site.register(Partner, PartnerAdmin)
admin.site.register(PartnerSettings, PartnerSettingsAdmin)
