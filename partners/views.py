import re
from django.shortcuts import render, redirect, get_object_or_404
from django.views.generic import TemplateView, ListView, DetailView, FormView, View
from django.contrib.auth.views import LoginView, LogoutView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.mail import send_mail
from django.conf import settings
from django.http import HttpResponseForbidden
from django.urls import reverse_lazy, reverse
from django.utils import timezone

from catalog.models import Category, Product
from .models import PartnerRequest, Partner, PartnerSettings, PartnerOrder, PartnerOrderItem
from .forms import PartnerRequestForm, PartnerLoginForm, PartnerProfileForm, PartnerPasswordChangeForm
from django.http import JsonResponse
from django.db.models import Q
from django.core.mail import EmailMessage
from django.utils.html import escape


def get_partner_or_none(user):
    """Безопасно получить профиль партнёра или None для админов."""
    if hasattr(user, 'partner_profile'):
        return user.partner_profile
    return None


class PartnerRequiredMixin(LoginRequiredMixin):
    """Миксин для проверки, что пользователь является активным партнёром или админом."""
    login_url = '/partners/login/'
    
    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return self.handle_no_permission()
        
        # Админы (is_staff) имеют доступ к партнёрскому разделу без проверки
        if request.user.is_staff:
            return super().dispatch(request, *args, **kwargs)
        
        # Проверяем, есть ли профиль партнёра
        if not hasattr(request.user, 'partner_profile'):
            messages.error(request, 'У вас нет доступа к разделу для партнёров.')
            return redirect('partners:wholesale')
        
        # Проверяем, активен ли партнёр
        if not request.user.partner_profile.is_active:
            messages.error(request, 'Ваш партнёрский аккаунт деактивирован. Обратитесь к менеджеру.')
            return redirect('partners:wholesale')
        
        return super().dispatch(request, *args, **kwargs)


class WholesaleView(TemplateView):
    """Главная страница раздела для партнёров."""
    template_name = 'partners/wholesale.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['partner_settings'] = PartnerSettings.get_settings()
        context['is_partner'] = self.is_partner()
        
        # Несколько товаров для превью (ТОЛЬКО из партнёрского каталога!)
        context['preview_products'] = Product.objects.filter(
            is_active=True,
            catalog_type='wholesale'
        ).select_related('category').prefetch_related('images')[:6]
        
        return context
    
    def is_partner(self):
        """Проверка, является ли текущий пользователь партнёром."""
        if not self.request.user.is_authenticated:
            return False
        return (
            hasattr(self.request.user, 'partner_profile') and 
            self.request.user.partner_profile.is_active
        )


class PartnerRegisterView(FormView):
    """Страница регистрации партнёра (подача заявки)."""
    template_name = 'partners/register.html'
    form_class = PartnerRequestForm
    success_url = reverse_lazy('partners:register_success')
    
    def form_valid(self, form):
        request_obj = form.save()
        
        # Отправляем уведомление менеджеру
        self.send_notification_to_manager(request_obj)
        
        return super().form_valid(form)
    
    def send_notification_to_manager(self, request_obj):
        """Отправка уведомления менеджеру о новой заявке."""
        try:
            partner_settings = PartnerSettings.get_settings()
            manager_email = partner_settings.manager_email or settings.MANAGER_EMAIL
            
            subject = f'Новая заявка на партнёрство от {request_obj.full_name}'
            message = f'''
Новая заявка на партнёрство:

ФИО: {request_obj.full_name}
Телефон: {request_obj.phone}
Email: {request_obj.email}
Город: {request_obj.city}
Комментарий: {request_obj.comment or 'Не указан'}

Дата заявки: {request_obj.created_at.strftime('%d.%m.%Y %H:%M')}

Перейти в админку для обработки:
{self.request.build_absolute_uri('/admin/partners/partnerrequest/')}
            '''
            
            send_mail(
                subject=subject,
                message=message,
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[manager_email],
                fail_silently=True,
            )
        except Exception as e:
            # Логируем ошибку, но не прерываем процесс
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f'Error sending partner notification email: {e}')


class PartnerRegisterSuccessView(TemplateView):
    """Страница успешной отправки заявки."""
    template_name = 'partners/register_success.html'


class PartnerLoginView(LoginView):
    """Страница авторизации партнёра."""
    template_name = 'partners/login.html'
    authentication_form = PartnerLoginForm
    redirect_authenticated_user = False  # Не перенаправлять авторизованных, покажем форму
    
    def get(self, request, *args, **kwargs):
        # Если пользователь уже авторизован — перенаправляем в каталог
        if request.user.is_authenticated:
            # Админы или активные партнёры — сразу в каталог
            if request.user.is_staff:
                return redirect('partners:catalog')
            if hasattr(request.user, 'partner_profile') and request.user.partner_profile.is_active:
                return redirect('partners:catalog')
        return super().get(request, *args, **kwargs)
    
    def get_success_url(self):
        return reverse('partners:catalog')
    
    def form_valid(self, form):
        # Проверяем, есть ли у пользователя профиль партнёра или он админ
        user = form.get_user()
        
        # Админы (is_staff) имеют доступ без профиля партнёра
        if user.is_staff:
            return super().form_valid(form)
        
        if not hasattr(user, 'partner_profile'):
            messages.error(
                self.request, 
                'У вас нет доступа к разделу для партнёров. Обратитесь к менеджеру.'
            )
            return self.form_invalid(form)
        
        if not user.partner_profile.is_active:
            messages.error(
                self.request, 
                'Ваш партнёрский аккаунт деактивирован. Обратитесь к менеджеру.'
            )
            return self.form_invalid(form)
        
        # Обновляем время последнего входа
        user.partner_profile.last_login = timezone.now()
        user.partner_profile.save(update_fields=['last_login'])
        
        return super().form_valid(form)


class PartnerLogoutView(LogoutView):
    """Выход из партнёрского кабинета."""
    next_page = 'partners:wholesale'


class PartnerProductView(PartnerRequiredMixin, DetailView):
    """Детальная страница товара для партнёра."""
    model = Product
    template_name = 'partners/product_detail.html'
    context_object_name = 'product'
    slug_field = 'slug'
    slug_url_kwarg = 'slug'
    
    def get_queryset(self):
        # ТОЛЬКО товары из партнёрского каталога!
        return Product.objects.filter(
            is_active=True,
            catalog_type='wholesale'
        ).select_related('category').prefetch_related('images')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['partner'] = get_partner_or_none(self.request.user)
        
        # Похожие товары (только из партнёрского каталога)
        product = self.object
        
        # Получаем характеристики и добавляем вольтаж, если он есть в применимости
        characteristics = product.get_characteristics_list()
        voltage = product.get_voltage_from_applicability()
        if voltage:
            # Проверяем, нет ли уже вольтажа в характеристиках
            has_voltage = any(key.lower() in ['вольтаж', 'voltage', 'напряжение'] for key, _ in characteristics)
            if not has_voltage:
                characteristics.append(('Напряжение', voltage))
        context['characteristics'] = characteristics
        
        # Похожие товары - ищем по кросс-номерам и категории
        similar_products = []
        from django.db.models import Q
        
        # Приоритет 1: Товары с таким же кросс-номером (могут быть из разных категорий!)
        # Это взаимозаменяемые/связанные детали
        if product.cross_numbers:
            cross_list = product.get_cross_numbers_list()
            if cross_list:
                query = Q()
                for cross in cross_list[:5]:  # Берём первые 5 кросс-номеров
                    if len(cross) >= 5:  # Минимум 5 символов для поиска
                        query |= Q(cross_numbers__icontains=cross)
                        query |= Q(article__iexact=cross)  # Артикул может совпадать с кросс-номером
                
                if query:
                    similar_products = list(Product.objects.filter(
                        query,
                        is_active=True,
                        catalog_type='wholesale'
                    ).exclude(pk=product.pk).select_related('category').prefetch_related('images')[:4])
        
        # Приоритет 2: Та же категория + та же применимость
        if len(similar_products) < 4 and product.category and product.applicability:
            current_applicability = product.get_applicability_list()
            if current_applicability:
                search_models = current_applicability[:3]
                query = Q()
                for model in search_models:
                    query |= Q(applicability__icontains=model)
                
                existing_ids = [p.pk for p in similar_products]
                more_products = Product.objects.filter(
                    query,
                    category=product.category,
                    is_active=True,
                    catalog_type='wholesale'
                ).exclude(pk=product.pk).exclude(pk__in=existing_ids).select_related('category').prefetch_related('images')[:4 - len(similar_products)]
                similar_products.extend(list(more_products))
        
        # Приоритет 3: Та же категория + тот же бренд
        if len(similar_products) < 4 and product.category and product.brand:
            existing_ids = [p.pk for p in similar_products]
            more_products = Product.objects.filter(
                category=product.category,
                brand__iexact=product.brand.strip(),
                is_active=True,
                catalog_type='wholesale'
            ).exclude(pk=product.pk).exclude(pk__in=existing_ids).select_related('category').prefetch_related('images')[:4 - len(similar_products)]
            similar_products.extend(list(more_products))
        
        # Приоритет 4: Просто та же категория
        if len(similar_products) < 4 and product.category:
            existing_ids = [p.pk for p in similar_products]
            more_products = Product.objects.filter(
                category=product.category,
                is_active=True,
                catalog_type='wholesale'
            ).exclude(pk=product.pk).exclude(pk__in=existing_ids).select_related('category').prefetch_related('images')[:4 - len(similar_products)]
            similar_products.extend(list(more_products))
        
        context['similar_products'] = similar_products[:4]
        
        return context


class PartnerProfileView(PartnerRequiredMixin, TemplateView):
    """Профиль партнёра."""
    template_name = 'partners/profile.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['partner'] = get_partner_or_none(self.request.user)
        return context


class PartnerProfileEditView(PartnerRequiredMixin, FormView):
    """Редактирование профиля партнёра."""
    template_name = 'partners/profile_edit.html'
    form_class = PartnerProfileForm
    success_url = reverse_lazy('partners:profile')
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['instance'] = self.request.user.partner_profile
        return kwargs
    
    def form_valid(self, form):
        form.save()
        messages.success(self.request, 'Профиль успешно обновлён.')
        return super().form_valid(form)


class PartnerPasswordChangeView(PartnerRequiredMixin, FormView):
    """Смена пароля партнёра."""
    template_name = 'partners/password_change.html'
    form_class = PartnerPasswordChangeForm
    success_url = reverse_lazy('partners:profile')
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs
    
    def form_valid(self, form):
        form.save()
        messages.success(self.request, 'Пароль успешно изменён.')
        return super().form_valid(form)


# Публичный каталог (без авторизации, но без цен)
class PublicPartnerCatalogView(ListView):
    """Публичный каталог для партнёров (без цен и наличия)."""
    template_name = 'partners/public_catalog.html'
    context_object_name = 'products'
    paginate_by = 24
    
    def get_queryset(self):
        # ТОЛЬКО товары из партнёрского каталога!
        queryset = Product.objects.filter(
            is_active=True,
            catalog_type='wholesale'
        ).select_related('category').prefetch_related('images')
        
        # Фильтр по категории
        category_slug = self.kwargs.get('category_slug')
        if category_slug:
            category = get_object_or_404(Category, slug=category_slug, is_active=True)
            descendants = category.get_descendants(include_self=True)
            queryset = queryset.filter(category__in=descendants)
            self.current_category = category
        else:
            self.current_category = None
        
        # Поиск (по частичным совпадениям слов)
        search_query = self.request.GET.get('q', '').strip()
        if search_query:
            # Разбиваем запрос на отдельные слова (минимум 2 символа)
            query_words = [word.strip() for word in search_query.split() if len(word.strip()) >= 2]
            
            if not query_words:
                # Если слово слишком короткое, ищем весь запрос целиком
                query_words = [search_query]
            
            # Для каждого слова создаём условие поиска
            # Используем AND - товар должен содержать ВСЕ слова из запроса
            for word in query_words:
                # Названия товаров хранятся с большой буквы (кириллица)
                # Преобразуем первую букву в заглавную для поиска в name
                # Для кириллицы capitalize() работает правильно
                word_capitalize = word.capitalize() if word else word
                word_upper = word.upper()
                word_escaped = word.replace('\\', '\\\\').replace('(', '\\(').replace(')', '\\)').replace('[', '\\[').replace(']', '\\]')
                
                # Для name (название) - используем contains с преобразованием первой буквы в заглавную
                # Для остальных полей - используем icontains/iregex (работает для латиницы и цифр)
                word_q = (
                    # Название товара - ищем с заглавной буквы (как хранится в базе: "Крестовина")
                    # Это работает когда пользователь вводит "крестовина" -> преобразуется в "Крестовина"
                    Q(name__contains=word_capitalize) |
                    # Также пробуем в верхнем регистре (на случай, если есть варианты "КРЕСТОВИНА")
                    Q(name__contains=word_upper) |
                    # И исходный регистр (на случай, если пользователь ввел с большой буквы)
                    Q(name__contains=word) |
                    # Остальные поля - регистронезависимый поиск
                    Q(article__icontains=word) |
                    Q(brand__icontains=word) |
                    Q(cross_numbers__icontains=word) |
                    Q(applicability__icontains=word) |
                    Q(description__icontains=word) |
                    Q(short_description__icontains=word) |
                    # Также пробуем iregex для надежности
                    Q(article__iregex=word_escaped) |
                    Q(brand__iregex=word_escaped) |
                    Q(cross_numbers__iregex=word_escaped) |
                    Q(applicability__iregex=word_escaped) |
                    Q(description__iregex=word_escaped) |
                    Q(short_description__iregex=word_escaped)
                )
                queryset = queryset.filter(word_q)
            
            self.search_query = search_query
        else:
            self.search_query = ''
        
        return queryset.order_by('name')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['categories'] = Category.objects.filter(
            parent=None, 
            is_active=True
        ).order_by('order', 'name')
        context['current_category'] = getattr(self, 'current_category', None)
        context['search_query'] = getattr(self, 'search_query', '')
        context['partner_settings'] = PartnerSettings.get_settings()
        return context


class PublicPartnerProductView(DetailView):
    """Публичная страница товара для неавторизованных."""
    model = Product
    template_name = 'partners/public_product_detail.html'
    context_object_name = 'product'
    slug_field = 'slug'
    slug_url_kwarg = 'slug'
    
    def get_queryset(self):
        # ТОЛЬКО товары из партнёрского каталога!
        return Product.objects.filter(
            is_active=True,
            catalog_type='wholesale'
        ).select_related('category').prefetch_related('images')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['partner_settings'] = PartnerSettings.get_settings()
        
        # Получаем характеристики и добавляем вольтаж, если он есть в применимости
        product = self.object
        characteristics = product.get_characteristics_list()
        voltage = product.get_voltage_from_applicability()
        if voltage:
            # Проверяем, нет ли уже вольтажа в характеристиках
            has_voltage = any(key.lower() in ['вольтаж', 'voltage', 'напряжение'] for key, _ in characteristics)
            if not has_voltage:
                characteristics.append(('Напряжение', voltage))
        context['characteristics'] = characteristics
        
        return context


# ============================================
# Заказы партнёров
# ============================================

def get_partner_cart(request):
    """Получить корзину партнёра из сессии."""
    # Админы тоже могут использовать корзину для тестирования
    if not request.user.is_authenticated:
        return {}
    if not hasattr(request.user, 'partner_profile') and not request.user.is_staff:
        return {}
    cart = request.session.get('partner_cart', {})
    return cart


def set_partner_cart(request, cart):
    """Сохранить корзину партнёра в сессию."""
    request.session['partner_cart'] = cart
    request.session.modified = True


class PartnerCatalogView(PartnerRequiredMixin, ListView):
    """Каталог товаров для партнёров."""
    template_name = 'partners/catalog.html'
    context_object_name = 'products'
    paginate_by = 24
    
    def get_queryset(self):
        # ТОЛЬКО товары из партнёрского каталога!
        queryset = Product.objects.filter(
            is_active=True,
            catalog_type='wholesale'
        ).select_related('category').prefetch_related('images')
        
        # Фильтр по категории
        category_slug = self.kwargs.get('category_slug')
        if category_slug:
            category = get_object_or_404(Category, slug=category_slug, is_active=True)
            # Включаем товары из подкатегорий
            descendants = category.get_descendants(include_self=True)
            queryset = queryset.filter(category__in=descendants)
            self.current_category = category
        else:
            self.current_category = None
        
        # Поиск (по частичным совпадениям слов)
        search_query = self.request.GET.get('q', '').strip()
        if search_query:
            # Разбиваем запрос на отдельные слова (минимум 2 символа)
            query_words = [word.strip() for word in search_query.split() if len(word.strip()) >= 2]
            
            if not query_words:
                # Если слово слишком короткое, ищем весь запрос целиком
                query_words = [search_query]
            
            # Для каждого слова создаём условие поиска
            # Используем AND - товар должен содержать ВСЕ слова из запроса
            for word in query_words:
                # Названия товаров хранятся с большой буквы (кириллица)
                # Преобразуем первую букву в заглавную для поиска в name
                # Для кириллицы capitalize() работает правильно
                word_capitalize = word.capitalize() if word else word
                word_upper = word.upper()
                word_escaped = word.replace('\\', '\\\\').replace('(', '\\(').replace(')', '\\)').replace('[', '\\[').replace(']', '\\]')
                
                # Для name (название) - используем contains с преобразованием первой буквы в заглавную
                # Для остальных полей - используем icontains/iregex (работает для латиницы и цифр)
                word_q = (
                    # Название товара - ищем с заглавной буквы (как хранится в базе: "Крестовина")
                    # Это работает когда пользователь вводит "крестовина" -> преобразуется в "Крестовина"
                    Q(name__contains=word_capitalize) |
                    # Также пробуем в верхнем регистре (на случай, если есть варианты "КРЕСТОВИНА")
                    Q(name__contains=word_upper) |
                    # И исходный регистр (на случай, если пользователь ввел с большой буквы)
                    Q(name__contains=word) |
                    # Остальные поля - регистронезависимый поиск
                    Q(article__icontains=word) |
                    Q(brand__icontains=word) |
                    Q(cross_numbers__icontains=word) |
                    Q(applicability__icontains=word) |
                    Q(description__icontains=word) |
                    Q(short_description__icontains=word) |
                    # Также пробуем iregex для надежности
                    Q(article__iregex=word_escaped) |
                    Q(brand__iregex=word_escaped) |
                    Q(cross_numbers__iregex=word_escaped) |
                    Q(applicability__iregex=word_escaped) |
                    Q(description__iregex=word_escaped) |
                    Q(short_description__iregex=word_escaped)
                )
                queryset = queryset.filter(word_q)
            
            self.search_query = search_query
        else:
            self.search_query = ''
        
        # Сортировка по умолчанию - по алфавиту
        sort = self.request.GET.get('sort', 'name')
        if sort == 'price_asc':
            queryset = queryset.order_by('wholesale_price', 'price')
        elif sort == 'price_desc':
            queryset = queryset.order_by('-wholesale_price', '-price')
        elif sort == 'name':
            queryset = queryset.order_by('name')
        elif sort == '-created_at':
            queryset = queryset.order_by('-created_at')
        else:
            queryset = queryset.order_by('name')
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['categories'] = Category.objects.filter(
            parent=None, 
            is_active=True
        ).order_by('order', 'name')
        context['current_category'] = getattr(self, 'current_category', None)
        context['search_query'] = getattr(self, 'search_query', '')
        context['current_sort'] = self.request.GET.get('sort', 'name')
        context['partner'] = get_partner_or_none(self.request.user)
        
        # Получаем корзину для отображения количества
        cart = get_partner_cart(self.request)
        context['cart_count'] = sum(item['quantity'] for item in cart.values())
        
        return context


@login_required
def partner_cart_add(request, product_id):
    """Добавить товар в корзину партнёра."""
    if not hasattr(request.user, 'partner_profile') and not request.user.is_staff:
        return JsonResponse({'success': False, 'error': 'Недостаточно прав'}, status=403)
    
    if request.method == 'POST':
        product = get_object_or_404(
            Product, 
            id=product_id, 
            is_active=True,
            catalog_type='wholesale'
        )
        cart = get_partner_cart(request)
        
        product_id_str = str(product_id)
        quantity = int(request.POST.get('quantity', 1))
        
        if product_id_str in cart:
            cart[product_id_str]['quantity'] += quantity
        else:
            # Используем оптовую цену, если есть
            price = float(product.wholesale_price or product.price)
            cart[product_id_str] = {
                'quantity': quantity,
                'price': str(price),
            }
        
        set_partner_cart(request, cart)
        
        cart_count = sum(item['quantity'] for item in cart.values())
        
        return JsonResponse({
            'success': True,
            'cart_count': cart_count,
            'message': f'Товар "{product.name}" добавлен в корзину'
        })
    
    return JsonResponse({'success': False, 'error': 'Метод не разрешён'}, status=405)


@login_required
def partner_cart_update(request, product_id):
    """Изменить количество товара в корзине партнёра."""
    if not hasattr(request.user, 'partner_profile') and not request.user.is_staff:
        return JsonResponse({'success': False, 'error': 'Недостаточно прав'}, status=403)
    
    if request.method == 'POST':
        cart = get_partner_cart(request)
        product_id_str = str(product_id)
        quantity = int(request.POST.get('quantity', 1))
        
        if product_id_str in cart:
            if quantity > 0:
                cart[product_id_str]['quantity'] = quantity
            else:
                del cart[product_id_str]
            
            set_partner_cart(request, cart)
        
        cart_count = sum(item['quantity'] for item in cart.values())
        return JsonResponse({'success': True, 'cart_count': cart_count})
    
    return JsonResponse({'success': False, 'error': 'Метод не разрешён'}, status=405)


@login_required
def partner_cart_count(request):
    """Получить количество товаров в корзине партнёра."""
    if not hasattr(request.user, 'partner_profile') and not request.user.is_staff:
        return JsonResponse({'count': 0})
    
    cart = get_partner_cart(request)
    count = sum(item['quantity'] for item in cart.values())
    return JsonResponse({'count': count})


class PartnerOrdersView(PartnerRequiredMixin, ListView):
    """Список заказов партнёра."""
    template_name = 'partners/orders.html'
    context_object_name = 'orders'
    paginate_by = 20
    
    def get_queryset(self):
        partner = get_partner_or_none(self.request.user)
        # Админы видят все заказы, партнёры — только свои
        if partner:
            queryset = PartnerOrder.objects.filter(partner=partner).prefetch_related('items__product')
        elif self.request.user.is_staff:
            queryset = PartnerOrder.objects.all().prefetch_related('items__product')
        else:
            queryset = PartnerOrder.objects.none()
        
        # Фильтр по дате
        date_from = self.request.GET.get('date_from', '').strip()
        date_to = self.request.GET.get('date_to', '').strip()
        
        if date_from:
            try:
                from datetime import datetime
                date_from_obj = datetime.strptime(date_from, '%Y-%m-%d')
                queryset = queryset.filter(created_at__gte=date_from_obj)
            except ValueError:
                pass
        
        if date_to:
            try:
                from datetime import datetime
                date_to_obj = datetime.strptime(date_to, '%Y-%m-%d')
                queryset = queryset.filter(created_at__lte=date_to_obj)
            except ValueError:
                pass
        
        # Поиск по названию, OEM, артикулу
        search_query = self.request.GET.get('q', '').strip()
        if search_query:
            queryset = queryset.filter(
                Q(items__product__name__icontains=search_query) |
                Q(items__product__article__icontains=search_query) |
                Q(items__product__brand__icontains=search_query) |
                Q(items__product__cross_numbers__icontains=search_query)
            ).distinct()
        
        return queryset.order_by('-created_at')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['partner'] = get_partner_or_none(self.request.user)
        context['date_from'] = self.request.GET.get('date_from', '')
        context['date_to'] = self.request.GET.get('date_to', '')
        context['search_query'] = self.request.GET.get('q', '')
        return context


@login_required
def partner_order_create(request):
    """Создать заказ из корзины партнёра."""
    # Админы могут просматривать, но не создавать заказы без профиля партнёра
    if not hasattr(request.user, 'partner_profile'):
        if request.user.is_staff:
            messages.warning(request, 'Для оформления заказов необходим профиль партнёра. Вы вошли как администратор.')
        else:
            messages.error(request, 'Недостаточно прав')
        return redirect('partners:catalog')
    
    cart = get_partner_cart(request)
    
    if not cart:
        messages.warning(request, 'Ваша корзина пуста')
        return redirect('partners:catalog')
    
    partner = request.user.partner_profile
    
    # Создаём заказ со статусом draft (черновик)
    order = PartnerOrder.objects.create(
        partner=partner,
        status='draft'
    )
    
    # Добавляем товары
    for product_id, item_data in cart.items():
        try:
            product = Product.objects.get(id=int(product_id), is_active=True)
            quantity = item_data['quantity']
            price = float(item_data['price'])
            
            PartnerOrderItem.objects.create(
                order=order,
                product=product,
                quantity=quantity,
                price=price,
            )
        except Product.DoesNotExist:
            continue
    
    # Очищаем корзину
    set_partner_cart(request, {})
    
    # Email отправляется только при подтверждении заказа (статус pending)
    # Здесь заказ создаётся со статусом draft (черновик)
    
    messages.success(request, f'Заказ #{order.id} успешно создан. Подтвердите заказ на странице "Мои заказы"')
    return redirect('partners:orders')


@login_required
def partner_order_item_remove(request, order_id, item_id):
    """Удалить товар из заказа."""
    if not hasattr(request.user, 'partner_profile'):
        if request.user.is_staff:
            messages.warning(request, 'Для работы с заказами необходим профиль партнёра.')
        else:
            messages.error(request, 'Недостаточно прав')
        return redirect('partners:orders')
    
    order = get_object_or_404(
        PartnerOrder,
        id=order_id,
        partner=request.user.partner_profile,
        status__in=['draft', 'pending']
    )
    
    item = get_object_or_404(PartnerOrderItem, id=item_id, order=order)
    item.delete()
    
    # Если в заказе не осталось товаров, удаляем заказ
    if order.items.count() == 0:
        order.delete()
        messages.success(request, 'Заказ удалён (в нём не осталось товаров)')
    else:
        messages.success(request, 'Товар удалён из заказа')
    
    return redirect('partners:orders')


@login_required
def partner_order_confirm(request, order_id):
    """Подтвердить заказ (изменить статус на pending)."""
    if not hasattr(request.user, 'partner_profile'):
        if request.user.is_staff:
            messages.warning(request, 'Для работы с заказами необходим профиль партнёра.')
        else:
            messages.error(request, 'Недостаточно прав')
        return redirect('partners:orders')
    
    order = get_object_or_404(
        PartnerOrder,
        id=order_id,
        partner=request.user.partner_profile,
        status='draft'
    )
    
    if order.items.count() == 0:
        messages.warning(request, 'Нельзя подтвердить пустой заказ')
        return redirect('partners:orders')
    
    if request.method == 'POST':
        # Сохраняем комментарий, если он есть
        comment = request.POST.get('comment', '').strip()
        order.comment = comment
        order.status = 'pending'
        order.save()
        
        # Уменьшаем остатки товаров
        for item in order.items.all():
            product = item.product
            if product.quantity >= item.quantity:
                product.quantity -= item.quantity
                # Обновляем статус наличия, если остаток стал 0
                if product.quantity == 0:
                    product.availability = 'out_of_stock'
                elif product.availability == 'out_of_stock' and product.quantity > 0:
                    product.availability = 'in_stock'
                product.save(update_fields=['quantity', 'availability'])
            else:
                # Если остатка недостаточно, уменьшаем до 0
                product.quantity = 0
                product.availability = 'out_of_stock'
                product.save(update_fields=['quantity', 'availability'])
        
        # Отправляем email менеджеру о подтверждённом заказе
        send_partner_order_email(order, request)
        
        messages.success(request, f'Заказ #{order.id} подтверждён и отправлен на обработку')
        return redirect('partners:orders')
    else:
        # GET запрос - просто меняем статус без комментария
        order.status = 'pending'
        order.save()
        
        # Уменьшаем остатки товаров
        for item in order.items.all():
            product = item.product
            if product.quantity >= item.quantity:
                product.quantity -= item.quantity
                # Обновляем статус наличия, если остаток стал 0
                if product.quantity == 0:
                    product.availability = 'out_of_stock'
                elif product.availability == 'out_of_stock' and product.quantity > 0:
                    product.availability = 'in_stock'
                product.save(update_fields=['quantity', 'availability'])
            else:
                # Если остатка недостаточно, уменьшаем до 0
                product.quantity = 0
                product.availability = 'out_of_stock'
                product.save(update_fields=['quantity', 'availability'])
        
        # Отправляем email менеджеру о подтверждённом заказе
        send_partner_order_email(order, request)
        
        messages.success(request, f'Заказ #{order.id} подтверждён и отправлен на обработку')
        return redirect('partners:orders')


def send_partner_order_email(order, request=None):
    """Отправка email менеджеру о новом заказе партнёра с прикреплённым файлом XLS."""
    from django.utils import timezone
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from io import BytesIO
    
    # Формируем URL для админки
    if request:
        admin_url = request.build_absolute_uri(f'/admin/partners/partnerorder/{order.id}/')
    else:
        # Используем настройки сайта или дефолтный URL
        site_url = getattr(settings, 'SITE_URL', 'http://localhost:8000')
        admin_url = f'{site_url}/admin/partners/partnerorder/{order.id}/'
    
    # Создаём Excel файл
    wb = Workbook()
    ws = wb.active
    ws.title = f'Заказ #{order.id}'
    
    # Заголовки (товары слева, контакты справа)
    headers = ['№ заказа', 'Дата', 'Статус', 'Товар', 'Артикул', 'Бренд', 'Количество', 'Цена', 'Сумма', 'Партнёр', 'Компания', 'Телефон', 'Email', 'Комментарий']
    ws.append(headers)
    
    # Стили для заголовков
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Данные (товары слева, контакты справа)
    for item in order.items.all():
        ws.append([
            order.id,
            order.created_at.strftime('%d.%m.%Y %H:%M'),
            order.get_status_display(),
            item.product.name,
            item.product.article or '',
            item.product.brand or '',
            item.quantity,
            float(item.price),
            float(item.get_total()),
            order.partner.full_name,
            order.partner.company_name or '',
            order.partner.phone,
            order.partner.email,
            order.comment or '',
        ])
    
    # Добавляем итоговую строку
    # Колонки: 0-№ заказа, 1-Дата, 2-Статус, 3-Товар, 4-Артикул, 5-Бренд, 6-Количество, 7-Цена, 8-Сумма, 9-Партнёр, 10-Компания, 11-Телефон, 12-Email, 13-Комментарий
    ws.append([])
    ws.append(['Итого:', '', '', '', '', '', order.get_total_quantity(), '', order.get_total_price(), '', '', '', '', ''])
    
    # Автоширина колонок
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Сохраняем в BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    # Используем локальное время
    local_time = timezone.localtime(order.created_at)
    
    partner = order.partner
    
    subject = f'[ЗАКАЗ ПАРТНЁРА] Заказ #{order.id} от {partner.full_name}'
    
    # Формируем список товаров
    items_list = []
    for item in order.items.all():
        items_list.append(
            f"- {item.product.name} (Кросс-номер: {item.product.article or 'не указан'})\n"
            f"  Количество: {item.quantity} шт.\n"
            f"  Цена: {item.price} руб. за шт.\n"
            f"  Сумма: {item.get_total()} руб."
        )
    
    items_text = '\n\n'.join(items_list) if items_list else 'Нет товаров'
    
    # Формируем HTML версию письма
    html_message = f"""
    <html>
    <head>
        <meta charset="UTF-8">
        <style>
            body {{ font-family: Arial, sans-serif; line-height: 1.6; color: #333; }}
            .header {{ background-color: #dc2626; color: white; padding: 20px; }}
            .content {{ padding: 20px; }}
            .order-info {{ background-color: #f7fafc; padding: 15px; margin: 10px 0; border-radius: 5px; }}
            .items {{ margin: 15px 0; }}
            .item {{ background-color: #edf2f7; padding: 10px; margin: 5px 0; border-left: 3px solid #dc2626; }}
            .total {{ font-size: 1.2em; font-weight: bold; color: #dc2626; margin-top: 15px; }}
            .footer {{ margin-top: 20px; padding-top: 20px; border-top: 1px solid #e2e8f0; color: #718096; font-size: 0.9em; }}
            .status {{ display: inline-block; padding: 5px 10px; border-radius: 5px; font-weight: bold; }}
            .status-pending {{ background-color: #ffc107; color: #000; }}
            .status-draft {{ background-color: #6c757d; color: #fff; }}
        </style>
    </head>
    <body>
        <div class="header">
            <h2>Новый заказ от партнёра!</h2>
        </div>
        <div class="content">
            <div class="order-info">
                <h3>Информация о заказе</h3>
                <p><strong>Номер заказа:</strong> #{order.id}</p>
                <p><strong>Дата:</strong> {local_time.strftime('%d.%m.%Y %H:%M')}</p>
                <p><strong>Статус:</strong> <span class="status status-{order.status}">{order.get_status_display()}</span></p>
            </div>
            
            <div class="order-info">
                <h3>Данные партнёра</h3>
                <p><strong>ФИО:</strong> {escape(partner.full_name)}</p>
                <p><strong>Компания:</strong> {escape(partner.company_name) if partner.company_name else 'не указана'}</p>
                <p><strong>Телефон:</strong> <a href="tel:{escape(partner.phone)}">{escape(partner.phone)}</a></p>
                <p><strong>Email:</strong> {escape(partner.email)}</p>
                <p><strong>Город:</strong> {escape(partner.city)}</p>
            </div>
            
            <div class="items">
                <h3>Товары в заказе:</h3>
                {''.join([f'''
                <div class="item">
                    <p><strong>{escape(item.product.name)}</strong></p>
                    <p>Кросс-номер: {escape(item.product.article) if item.product.article else 'не указан'}</p>
                    <p>Бренд: {escape(item.product.brand) if item.product.brand else 'не указан'}</p>
                    <p>Количество: {item.quantity} шт. × {item.price} руб. = {item.get_total()} руб.</p>
                </div>
                ''' for item in order.items.all()])}
            </div>
            
            <div class="total">
                Итоговая сумма: {order.get_total_price()} руб.
            </div>
            
            {f'<div class="order-info"><h3>Комментарий партнёра</h3><p>{escape(order.comment)}</p></div>' if order.comment else ''}
            
            <div class="footer">
                <p><a href="{admin_url}">Перейти к заказу в админке</a></p>
                <p>Это автоматическое уведомление от сайта {getattr(settings, 'SITE_NAME', 'Onesimus')}</p>
            </div>
        </div>
    </body>
    </html>
    """
    
    # Текстовая версия
    text_message = f"""
Новый заказ от партнёра!

Номер заказа: #{order.id}
Дата: {local_time.strftime('%d.%m.%Y %H:%M')}
Статус: {order.get_status_display()}

Данные партнёра:
ФИО: {partner.full_name}
Компания: {partner.company_name or 'не указана'}
Телефон: {partner.phone}
Email: {partner.email}
Город: {partner.city}

Товары в заказе:
{items_text}

Итоговая сумма: {order.get_total_price()} руб.

{f'Комментарий партнёра: {order.comment}' if order.comment else ''}

Ссылка на заказ в админке:
{admin_url}

---
Это автоматическое уведомление от сайта {getattr(settings, 'SITE_NAME', 'Onesimus')}
"""
    
    # Получаем email менеджера из настроек партнёрского раздела
    partner_settings = PartnerSettings.get_settings()
    manager_email = partner_settings.manager_email or getattr(settings, 'MANAGER_EMAIL', settings.DEFAULT_FROM_EMAIL)
    
    try:
        email = EmailMessage(
            subject=subject,
            body=text_message,
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=[manager_email],
        )
        email.content_subtype = "html"
        email.body = html_message
        
        # Прикрепляем Excel файл
        from datetime import datetime
        # Формируем имя файла с именем партнёра и датой
        partner_name_safe = partner.full_name.replace(' ', '_').replace('/', '_').replace('\\', '_')[:30]
        date_str = datetime.now().strftime("%Y%m%d")
        filename = f'Заказ_{order.id}_{partner_name_safe}_{date_str}.xlsx'
        email.attach(filename, excel_file.read(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
        # Добавляем специальные заголовки
        email.extra_headers = {
            'X-Mail-Category': 'PartnerOrder',
            'X-Priority': '1',
            'X-Auto-Response-Suppress': 'All',
            'X-Order-ID': str(order.id),
            'X-Partner-ID': str(partner.id),
            'X-Site-Name': getattr(settings, 'SITE_NAME', 'Onesimus'),
        }
        
        email.send(fail_silently=False)
    except Exception as e:
        # Логируем ошибку, но не прерываем процесс
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f'Error sending partner order email: {e}')


@login_required
def partner_orders_export_xls(request):
    """Экспорт заказов партнёра в XLS."""
    if not hasattr(request.user, 'partner_profile') and not request.user.is_staff:
        return HttpResponseForbidden()
    
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from datetime import datetime
    
    partner = get_partner_or_none(request.user)
    
    # Фильтры - админы видят все заказы, партнёры только свои
    if partner:
        queryset = PartnerOrder.objects.filter(partner=partner).prefetch_related('items__product')
    elif request.user.is_staff:
        queryset = PartnerOrder.objects.all().prefetch_related('items__product')
    
    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()
    search_query = request.GET.get('q', '').strip()
    
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d')
            queryset = queryset.filter(created_at__gte=date_from_obj)
        except ValueError:
            pass
    
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d')
            queryset = queryset.filter(created_at__lte=date_to_obj)
        except ValueError:
            pass
    
    if search_query:
        queryset = queryset.filter(
            Q(items__product__name__icontains=search_query) |
            Q(items__product__article__icontains=search_query) |
            Q(items__product__brand__icontains=search_query)
        ).distinct()
    
    queryset = queryset.order_by('-created_at')
    
    # Создаём Excel файл
    wb = Workbook()
    ws = wb.active
    ws.title = 'Заказы'
    
    # Заголовки
    headers = ['№ заказа', 'Дата', 'Статус', 'Товар', 'Артикул', 'Бренд', 'Количество', 'Цена', 'Сумма']
    ws.append(headers)
    
    # Стили для заголовков
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Данные
    for order in queryset:
        for item in order.items.all():
            ws.append([
                order.id,
                order.created_at.strftime('%d.%m.%Y %H:%M'),
                order.get_status_display(),
                item.product.name,
                item.product.article or '',
                item.product.brand or '',
                item.quantity,
                float(item.price),
                float(item.get_total()),
            ])
    
    # Автоширина колонок
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Сохраняем в HttpResponse
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="orders_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    wb.save(response)
    return response
